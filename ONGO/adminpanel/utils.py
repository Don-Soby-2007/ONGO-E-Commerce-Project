import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Count, Sum, Q
from decimal import Decimal
from datetime import datetime
from order.models import OrderItem
from returns.models import Return
from django.utils import timezone
from weasyprint import HTML
from django.template.loader import render_to_string


def safe_decimal(value):
    return value if value is not None else Decimal('0')


def generate_analytics_excel(request, queryset):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Analytics"

    # --- Styling Definitions ---
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    section_header_font = Font(bold=True, size=14, color="000000")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    currency_format = '#,##0.00'
    date_format = 'yyyy-mm-dd'

    def apply_table_style(start_row, start_col, end_row, end_col, has_header=True):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if has_header and row == start_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align

    current_row = 1

    # 1 Header & Meta Info
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="SALES ANALYTICS REPORT")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = center_align

    current_row += 2
    # Determine Filter Label
    date_filter = request.GET.get('date_filter', 'all')
    filter_label = f"Filter: {date_filter.upper()}"
    if date_filter == 'custom':
        filter_label += f" ({request.GET.get('start_date')} to {request.GET.get('end_date')})"

    print(f"DEBUG - Date Filter Value: {repr(date_filter)}")

    ws.cell(row=current_row, column=1, value=f"Report Period: {filter_label}")
    ws.cell(row=current_row, column=3, value=f"Generated: {datetime.now().strftime('%b %d, %Y, %I:%M %p')}")
    current_row += 2

    # 2. Executive Summary (KPIs)

    order_stats = queryset.aggregate(
        total_orders=Count('id'),
        gross_revenue=Sum('sub_total'),
        discount_amount=Sum('promotional_discount'),
        coupon_discount_amount=Sum('coupon_discount'),
    )

    total_orders = safe_decimal(order_stats['total_orders'])
    gross_revenue = safe_decimal(order_stats['gross_revenue'])
    discount_amount = safe_decimal(order_stats['discount_amount'])
    coupon_discount_amount = safe_decimal(order_stats['coupon_discount_amount'])
    overall_discount = discount_amount + coupon_discount_amount
    net_revenue = gross_revenue - overall_discount

    ws.cell(row=current_row, column=1, value="EXECUTIVE SUMMARY")
    ws.cell(row=current_row, column=1).font = section_header_font
    current_row += 1

    # KPI Table Headers
    headers = ["Metric", "Value"]
    # Change vs Prev is hard without historical data,keeping structure
    for col, h in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=h)

    # KPI Data
    kpi_data = [
        ["Total Sales", net_revenue],
        ["Total Orders", total_orders],
        ["Total Discount", overall_discount],
        ["Gross Revenue", gross_revenue],
    ]

    start_table_row = current_row + 1
    for i, row_data in enumerate(kpi_data):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_table_row + i, column=col, value=val)
            if col == 2 and isinstance(val, Decimal):
                cell.number_format = currency_format

    apply_table_style(current_row, 1, start_table_row + len(kpi_data) - 1, 2)
    current_row = start_table_row + len(kpi_data) + 2

    # 3. Return & Refund Analysis

    try:
        returns_data = Return.objects.filter(
            order__in=queryset
        ).aggregate(
            total_returns=Count('id'),
            accepted_returns=Count('id', filter=Q(status='accepted')),
            total_refunded=Sum('refund_amount')
        )

        ws.cell(row=current_row, column=1, value="RETURN & REFUND ANALYSIS")
        ws.cell(row=current_row, column=1).font = section_header_font
        current_row += 1

        ret_headers = ["Total Returns", "Accepted Returns", "Total Refunded Amount"]
        ret_values = [
            safe_decimal(returns_data['total_returns']),
            safe_decimal(returns_data['accepted_returns']),
            safe_decimal(returns_data['total_refunded'])
        ]

        for col, h in enumerate(ret_headers, 1):
            ws.cell(row=current_row, column=col, value=h)

        current_row += 1
        for col, val in enumerate(ret_values, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            if col == 3:
                cell.number_format = currency_format

        apply_table_style(current_row - 1, 1, current_row, 3)
        current_row += 2
    except ImportError:
        current_row += 1

    # 4. Payment Distribution
    payment_stats = queryset.values('payment_method').annotate(
        total_amount=Sum('total_amount'),
        order_count=Count('id')
    ).order_by('-total_amount')

    ws.cell(row=current_row, column=1, value="PAYMENT DISTRIBUTION")
    ws.cell(row=current_row, column=1).font = section_header_font
    current_row += 1

    pay_headers = ["Payment Method", "Total Amount", "Order Count"]
    for col, h in enumerate(pay_headers, 1):
        ws.cell(row=current_row, column=col, value=h)

    start_pay_row = current_row + 1
    for i, item in enumerate(payment_stats):
        ws.cell(row=start_pay_row + i, column=1, value=item['payment_method'] or 'Unknown')
        cell_amt = ws.cell(row=start_pay_row + i, column=2, value=safe_decimal(item['total_amount']))
        cell_amt.number_format = currency_format
        ws.cell(row=start_pay_row + i, column=3, value=item['order_count'])

    if payment_stats:
        apply_table_style(current_row, 1, start_pay_row + len(payment_stats) - 1, 3)
        current_row = start_pay_row + len(payment_stats) + 2
    else:
        current_row += 2

    # 5. Coupon Usage Analysis
    coupon_orders = queryset.filter(coupon__isnull=False)
    coupon_stats = coupon_orders.values(
        'coupon__coupon_code', 'coupon__discount_type'
    ).annotate(
        times_used=Count('id'),
        total_discount=Sum('coupon_discount')
    ).order_by('-total_discount')

    ws.cell(row=current_row, column=1, value="COUPON USAGE ANALYSIS")
    ws.cell(row=current_row, column=1).font = section_header_font
    current_row += 1

    coup_headers = ["Coupon Code", "Discount Type", "Usage Count", "Total Discount"]
    for col, h in enumerate(coup_headers, 1):
        ws.cell(row=current_row, column=col, value=h)

    start_coup_row = current_row + 1
    for i, item in enumerate(coupon_stats):
        ws.cell(row=start_coup_row + i, column=1, value=item['coupon__coupon_code'] or 'Unknown')
        ws.cell(row=start_coup_row + i, column=2, value=item['coupon__discount_type'] or 'Standard')
        ws.cell(row=start_coup_row + i, column=3, value=item['times_used'])
        cell_disc = ws.cell(row=start_coup_row + i, column=4, value=safe_decimal(item['total_discount']))
        cell_disc.number_format = currency_format

    if coupon_stats:
        apply_table_style(current_row, 1, start_coup_row + len(coupon_stats) - 1, 4)
        current_row = start_coup_row + len(coupon_stats) + 2
    else:
        current_row += 2

    # 6. Top 10 Products

    try:
        order_items = OrderItem.objects.filter(
            order__in=queryset
        ).select_related('product_variant__product').only(
            'product_name', 'price_at_purchase', 'quantity', 'line_discount',
            'order__sub_total', 'order__promotional_discount', 'order__coupon_discount',
            'product_variant__product__id'
        )

        product_map = {}
        for item in order_items:
            pid = item.product_variant.product.id if item.product_variant else None
            if not pid:
                continue

            if pid not in product_map:
                product_map[pid] = {
                    'name': item.product_name,
                    'gross_amount': Decimal('0'),
                    'discount': Decimal('0'),
                }

            item_price = safe_decimal((item.price_at_purchase * item.quantity))
            item_discount = safe_decimal(item.line_discount)

            product_map[pid]['gross_amount'] += item_price
            product_map[pid]['discount'] += item_discount

            if item.order.sub_total and item.order.sub_total > 0:
                order_total_discount = (safe_decimal(item.order.promotional_discount) +
                                        safe_decimal(item.order.coupon_discount))
                item_share = item_price / safe_decimal(item.order.sub_total)
                product_map[pid]['discount'] += item_share * order_total_discount

        product_breakdown = []
        for pid, data in product_map.items():
            data['net_revenue'] = data['gross_amount'] - data['discount']
            product_breakdown.append(data)

        product_breakdown.sort(key=lambda x: x['gross_amount'], reverse=True)
        top_products = product_breakdown[:10]

        print('Debug: top product', top_products)

        ws.cell(row=current_row, column=1, value="TOP 10 SELLING PRODUCTS")
        ws.cell(row=current_row, column=1).font = section_header_font
        current_row += 1

        prod_headers = ["Product Name", "Units Sold", "Gross Amount", "Discount", "Net Revenue"]

        product_map_excel = {}
        for item in order_items:
            pid = item.product_variant.product.id if item.product_variant else None
            if not pid:
                continue
            if pid not in product_map_excel:
                product_map_excel[pid] = {'name': item.product_name, 'qty': 0,
                                          'gross': Decimal('0'),
                                          'disc': Decimal('0')}

            product_map_excel[pid]['qty'] += item.quantity
            item_price = safe_decimal((item.price_at_purchase * item.quantity))
            product_map_excel[pid]['gross'] += item_price

            # Discount allocation logic same as view
            item_discount = safe_decimal(item.line_discount)
            product_map_excel[pid]['disc'] += item_discount
            if item.order.sub_total and item.order.sub_total > 0:
                order_total_discount = (safe_decimal(item.order.promotional_discount) +
                                        safe_decimal(item.order.coupon_discount))
                item_share = item_price / safe_decimal(item.order.sub_total)
                product_map_excel[pid]['disc'] += item_share * order_total_discount

        final_products = []
        for pid, data in product_map_excel.items():
            net = data['gross'] - data['disc']
            final_products.append({**data, 'net': net})
        final_products.sort(key=lambda x: x['gross'], reverse=True)
        top_products = final_products[:10]

        for col, h in enumerate(prod_headers, 1):
            ws.cell(row=current_row, column=col, value=h)

        start_prod_row = current_row + 1
        for i, prod in enumerate(top_products):
            ws.cell(row=start_prod_row + i, column=1, value=prod['name'])
            ws.cell(row=start_prod_row + i, column=2, value=prod['qty'])
            cell_gross = ws.cell(row=start_prod_row + i, column=3, value=prod['gross'])
            cell_gross.number_format = currency_format
            cell_disc = ws.cell(row=start_prod_row + i, column=4, value=prod['disc'])
            cell_disc.number_format = currency_format
            cell_net = ws.cell(row=start_prod_row + i, column=5, value=prod['net'])
            cell_net.number_format = currency_format

        if top_products:
            apply_table_style(current_row, 1, start_prod_row + len(top_products) - 1, 5)
            current_row = start_prod_row + len(top_products) + 2
        else:
            current_row += 2
    except ImportError:
        current_row += 1

    # 7. Detailed Order Log
    ws.cell(row=current_row, column=1, value="DETAILED ORDER LOG")
    ws.cell(row=current_row, column=1).font = section_header_font
    current_row += 1

    order_headers = ["Order ID", "Date", "Customer", "Email", "Items", "Amount", "Status"]
    for col, h in enumerate(order_headers, 1):
        ws.cell(row=current_row, column=col, value=h)

    start_order_row = current_row + 1

    for i, order in enumerate(queryset):
        ws.cell(row=start_order_row + i, column=1, value=str(order.order_id))
        cell_date = ws.cell(row=start_order_row + i, column=2,
                            value=order.created_at.date() if order.created_at else '')
        cell_date.number_format = date_format
        ws.cell(row=start_order_row + i, column=3,
                value=order.user.username)
        ws.cell(row=start_order_row + i, column=4, value=order.user.email)
        # Count items
        item_count = order.items.count() if hasattr(order, 'items') else 0
        ws.cell(row=start_order_row + i, column=5, value=item_count)
        cell_amt = ws.cell(row=start_order_row + i, column=6, value=order.total_amount or 0)
        cell_amt.number_format = currency_format
        ws.cell(row=start_order_row + i, column=7, value=order.status.upper())

    apply_table_style(current_row, 1, start_order_row + len(queryset) - 1, 7)

    # Column Width Adjustment
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Sales_Analytics_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def generate_analytics_pdf(request, queryset):

    # --- 1. Datas same as Remember: make a util function for DRY
    today = timezone.now().date()

    # KPIs
    order_stats = queryset.aggregate(
        total_orders=Count('id'),
        gross_revenue=Sum('sub_total'),
        discount_amount=Sum('promotional_discount'),
        coupon_discount_amount=Sum('coupon_discount'),
    )

    total_orders = safe_decimal(order_stats['total_orders'])
    gross_revenue = safe_decimal(order_stats['gross_revenue'])
    discount_amount = safe_decimal(order_stats['discount_amount'])
    coupon_discount_amount = safe_decimal(order_stats['coupon_discount_amount'])
    overall_discount = discount_amount + coupon_discount_amount
    net_revenue = gross_revenue - overall_discount

    # Returns
    try:
        returns_data = Return.objects.filter(
            order__in=queryset
        ).aggregate(
            total_returns=Count('id'),
            accepted_returns=Count('id', filter=Q(status='accepted')),
            total_refunded=Sum('refund_amount')
        )
        total_returns = safe_decimal(returns_data['total_returns'])
        accepted_returns = safe_decimal(returns_data['accepted_returns'])
        total_refunded = safe_decimal(returns_data['total_refunded'])
    except Exception:
        total_returns = accepted_returns = total_refunded = Decimal('0')

    # Payments
    payment_stats = queryset.values('payment_method').annotate(
        total_amount=Sum('total_amount'),
        order_count=Count('id')
    ).order_by('-total_amount')

    # Coupons
    coupon_orders = queryset.filter(coupon__isnull=False)
    coupon_stats = coupon_orders.values(
        'coupon__coupon_code', 'coupon__discount_type'
    ).annotate(
        times_used=Count('id'),
        total_discount=Sum('coupon_discount')
    ).order_by('-total_discount')

    # Top Products
    try:
        order_items = OrderItem.objects.filter(
            order__in=queryset
        ).select_related('product_variant__product')

        product_map = {}
        for item in order_items:
            pid = item.product_variant.product.id if item.product_variant else None
            if not pid:
                continue

            if pid not in product_map:
                product_map[pid] = {
                    'name': item.product_name,
                    'qty': 0,
                    'gross': Decimal('0'),
                    'disc': Decimal('0')
                }

            product_map[pid]['qty'] += item.quantity
            item_price = safe_decimal(item.price_at_purchase * item.quantity)
            product_map[pid]['gross'] += item_price

            item_discount = safe_decimal(item.line_discount)
            product_map[pid]['disc'] += item_discount

            if item.order.sub_total and item.order.sub_total > 0:
                order_total_discount = (safe_decimal(item.order.promotional_discount) +
                                        safe_decimal(item.order.coupon_discount))
                item_share = item_price / safe_decimal(item.order.sub_total)
                product_map[pid]['disc'] += item_share * order_total_discount

        product_breakdown = []
        for pid, data in product_map.items():
            data['net'] = data['gross'] - data['disc']
            product_breakdown.append(data)

        product_breakdown.sort(key=lambda x: x['gross'], reverse=True)
        top_products = product_breakdown[:10]
    except Exception:
        top_products = []

    # --- 2. Prepare Context Data ---
    date_filter = request.GET.get('date_filter', 'all')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    filter_label = date_filter.upper().replace('_', ' ')
    if date_filter == 'custom' and start_date and end_date:
        filter_label = f"Custom ({start_date} to {end_date})"

    context = {
        'report_date': today.strftime('%b %d, %Y, %I:%M %p'),
        'filter_label': filter_label,
        'net_revenue': net_revenue,
        'total_orders': total_orders,
        'overall_discount': overall_discount,
        'gross_revenue': gross_revenue,
        'total_returns': total_returns,
        'accepted_returns': accepted_returns,
        'total_refunded': total_refunded,
        'payment_stats': payment_stats,
        'coupon_stats': coupon_stats,
        'top_products': top_products,
        'orders': queryset.select_related('user').order_by('-created_at'),
    }

    print('DEBUG top_product : ', top_products)

    html_string = render_to_string('adminpanel/sales_report.html', context)

    pdf_file = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = f"Sales_Analytics_{today.strftime('%Y-%m-%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
